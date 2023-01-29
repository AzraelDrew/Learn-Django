from django.shortcuts import render, redirect, HttpResponse
# from django.core.validators import RegexValidator  # 正则验证
# from django.core.exceptions import ValidationError  # 规则验证报错
from app import models  # 数据库
from app.utils.pagination import Pagination  # 分页组件(实现)

# from django.core.files.uploadedfile import InMemoryUploadedFile

# Create your views here.


def index(request):
    return render(request, "index.html")


# 部门列表
def derpat_list(request):
    # 获取所有的部门
    queryset = models.Department.objects.all()
    page_object = Pagination(request, queryset, page_size=5)
    context = {
        "queryset": page_object.page_queryset,
        "page_string": page_object.html()
    }
    return render(request, "depart_list.html", context)


# 添加部门
def derpat_add(request):
    if request.method == "GET":
        return render(request, "depart_add.html")
    # 获取用户通过POST提交的数据
    title = request.POST.get("title")
    # 保存到数据库
    models.Department.objects.create(title=title)
    # 重定向页面
    return redirect("/depart/list/")


# 删除部门
def derpat_delete(request):
    # 获取ID
    nid = request.GET.get("nid")
    # 删除部门
    models.Department.objects.filter(id=nid).delete()
    # 重定向
    return redirect("/depart/list/")


# 修改部门
def derpat_edit(request, nid):
    if request.method == "GET":
        # 查询当前部门名称
        row_object = models.Department.objects.filter(id=nid).first()
        # print(row_object.id,row_object.title)
        return render(request, "depart_edit.html", {"title": row_object.title})
    if request.method == "POST":
        # 获取用户提交的部门名称
        # title = request.POST.get("title")
        title = request.POST["title"]
        # 更新部门名称
        models.Department.objects.filter(id=nid).update(title=title)
        return redirect("/depart/list/")


# # 用户列表
# def user_list(request):
#     # 获取所有的用户
#     queryset = models.UserInfo.objects.all()
#     # print(queryset)
#     # for obj in queryset:
#     #     print(obj.id,obj.name,obj.password,obj.age,obj.account,obj.create_time.strftime("%Y-%m-%d"),obj.gender,obj.get_gender_display(),obj.depart_id,obj.depart.title)

#     #     # print(obj.get_gender_display())     #  会将对应数值的对应值输出    格式  obj.get_字段名_display
#     #     # obj.depart_id    # 获取数据库中存储的值

#     # obj.depart.title   #  很据ID自动去关联的表中查找数据    obj.字段名.列名
#     page_object = Pagination(request, queryset, page_size=5)
#     context = {
#         "queryset": page_object.page_queryset,
#         "page_string": page_object.html()
#     }
#     return render(request, "user_list.html", context)


# 批量上传文件
def derpat_multi(request):
    # 获取用户上传的文件对象
    file_object = request.FILES.get("exc")
    # print(type(file_object))
    from openpyxl import load_workbook
    # 将文件对象传递给openpyxl  由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    # 循环每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return redirect("/depart/list/")